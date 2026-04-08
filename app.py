# ══════════════════════════════════════════════════════════════════════════════
#  GHG CARBON INVENTORY — McKinsey Style
#  Streamlit web application
#  Based on: GHG Protocol Corporate Standard
#
#  in this version:
#    - Upload an Excel workbook (matching the Lab1_dashboard.xlsx template)
#    - The app reads every input cell automatically using openpyxl
#    - All session_state variables are overwritten with the Excel values
#    - The Dashboard and all pages update instantly with the imported data
# ══════════════════════════════════════════════════════════════════════════════

# ── IMPORTS ───────────────────────────────────────────────────────────────────
import streamlit as st   # the web-app framework
import pandas as pd      # used for DataFrame tables and the CSV export
import json              # save / restore all inputs as a JSON file
import io                # in-memory byte stream — needed to read uploaded files
from datetime import datetime  # timestamp for the report

# openpyxl is the library that reads and writes .xlsx files.
# load_workbook() opens an Excel workbook from a file path or a byte stream.
# data_only=True tells it to return the last CALCULATED VALUE of each formula
# cell rather than the formula string itself (e.g. returns 143510 not "=K14+K25").
from openpyxl import load_workbook


# ── PAGE CONFIGURATION ────────────────────────────────────────────────────────
# Must be the very first Streamlit call. Controls browser tab, layout, sidebar.
st.set_page_config(
    page_title="GHG Carbon Inventory",
    page_icon="◈",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ══════════════════════════════════════════════════════════════════════════════
#  McKINSEY-STYLE CSS
#  Injected as raw HTML. Every rule is explained inline.
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
/* ── Google Fonts ─────────────────────────────────────────────────────────────
   Playfair Display = editorial serif for headings and KPI numbers
   Source Sans 3    = clean sans-serif for body text
   Source Code Pro  = monospace for numeric data and emission-factor chips     */
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:wght@400;600;700&family=Source+Sans+3:wght@300;400;500;600&family=Source+Code+Pro:wght@400;500&display=swap');

/* ── Design tokens (CSS variables) ───────────────────────────────────────────
   Define the McKinsey colour palette once here; reference everywhere below.   */
:root {
    --white:     #FFFFFF;
    --off-white: #F7F7F5;
    --rule:      #E0DDD8;
    --light-grey:#F0EEE9;
    --mid-grey:  #9A9591;
    --body:      #1A1A1A;
    --navy:      #002B5B;
    --accent:    #005EB8;
    --accent-lt: #E8F0FA;
    --teal:      #00857C;
    --amber:     #C9600A;
    --green:     #1A6E3C;
    --green-lt:  #E6F4EC;
    --rule-h:    2px solid #002B5B;
}

/* ── Global page background ────────────────────────────────────────────────── */
html, body,
[data-testid="stAppViewContainer"],
[data-testid="stMain"] {
    background: var(--white) !important;
    color: var(--body);
    font-family: 'Source Sans 3', sans-serif;
    font-size: 15px;
}

/* ── Sidebar — deep navy panel ─────────────────────────────────────────────── */
[data-testid="stSidebar"] { background: var(--navy) !important; border-right: none; }
[data-testid="stSidebar"] * { color: #C8D8EE !important; }
[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.15) !important; }
[data-testid="stSidebar"] [data-testid="stMetricValue"] {
    color: #FFFFFF !important;
    font-family: 'Source Code Pro', monospace !important;
    font-size: 1.05rem !important;
}
[data-testid="stSidebar"] [data-testid="stMetricLabel"] {
    color: #7DA3CC !important; font-size: 11px !important;
}

/* ── Headings ──────────────────────────────────────────────────────────────── */
h1 {
    font-family: 'Playfair Display', serif;
    font-weight: 700; font-size: 2rem; color: var(--navy);
    border-bottom: var(--rule-h); padding-bottom: 12px;
    margin-bottom: 4px; letter-spacing: -0.02em;
}
h2 {
    font-family: 'Playfair Display', serif;
    font-weight: 600; font-size: 1.2rem; color: var(--navy);
    margin-top: 28px; margin-bottom: 4px;
    border-bottom: 1px solid var(--rule); padding-bottom: 6px;
}

/* ── Eyebrow label (small uppercase tag above h1) ──────────────────────────── */
.eyebrow {
    font-size: 11px; font-weight: 600; text-transform: uppercase;
    letter-spacing: 0.15em; color: var(--accent);
    margin-bottom: 4px; display: block;
}

/* ── Page caption (subtitle beneath the page title) ────────────────────────── */
.page-caption { font-size: 13px; color: var(--mid-grey); margin-top: 2px; margin-bottom: 20px; }

/* ── KPI Card — the signature McKinsey metric block ────────────────────────────
   No box, no shadow. Just a thick navy top-border and Playfair serif number.  */
.mck-kpi { border-top: 3px solid var(--navy); padding: 16px 0 12px 0; }
.mck-kpi-label {
    font-size: 11px; font-weight: 600; text-transform: uppercase;
    letter-spacing: 0.12em; color: var(--mid-grey); margin-bottom: 6px;
}
.mck-kpi-value {
    font-family: 'Playfair Display', serif;
    font-size: 2.3rem; font-weight: 700; color: var(--navy);
    line-height: 1; margin-bottom: 4px;
}
.mck-kpi-unit  { font-size: 12px; color: var(--mid-grey); }
.mck-kpi-delta { font-size: 12px; font-weight: 600; margin-top: 4px; }
.delta-neg { color: var(--teal); }   /* emissions DOWN = good = teal  */
.delta-pos { color: var(--amber); }  /* emissions UP   = bad  = amber */

/* ── Insight box — highlighted callout with blue left-border ────────────────── */
.insight-box {
    background: var(--accent-lt); border-left: 4px solid var(--accent);
    padding: 12px 16px; margin: 12px 0; font-size: 13px; line-height: 1.5;
}
.insight-box strong { color: var(--navy); }

/* ── Success box — green variant for import confirmations ─────────────────────
   Same layout as insight-box but uses green to signal a successful action.    */
.success-box {
    background: var(--green-lt); border-left: 4px solid var(--green);
    padding: 12px 16px; margin: 12px 0; font-size: 13px; line-height: 1.6;
}
.success-box strong { color: var(--green); }
.success-box ul { margin: 8px 0 0 16px; padding: 0; }
.success-box li { margin-bottom: 3px; }

/* ── McKinsey data table ────────────────────────────────────────────────────── */
.mck-table { width:100%; border-collapse:collapse; font-size:13px; margin-top:8px; }
.mck-table thead tr { border-top:2px solid var(--navy); border-bottom:1px solid var(--navy); }
.mck-table thead th {
    padding:8px 12px; text-align:left;
    font-size:11px; font-weight:600; text-transform:uppercase;
    letter-spacing:0.08em; color:var(--navy);
}
.mck-table thead th.num { text-align:right; }
.mck-table tbody tr { border-bottom:1px solid var(--rule); }
.mck-table tbody tr:last-child { border-bottom:2px solid var(--navy); }
.mck-table tbody td { padding:8px 12px; }
.mck-table tbody td.num { text-align:right; font-family:'Source Code Pro',monospace; font-size:12px; }
.mck-table tr.total-row { background:var(--light-grey); }
.mck-table tr.total-row td { color:var(--navy); font-weight:600; }

/* ── Emission factor chip — monospace badge for showing factor values ──────── */
.ef-chip {
    display:inline-block; font-family:'Source Code Pro',monospace; font-size:11px;
    color:var(--accent); background:var(--accent-lt);
    padding:2px 7px; border-radius:2px; border:1px solid #C0D4EF;
}

/* ── Progress bar ─────────────────────────────────────────────────────────── */
.mck-progress-label { display:flex; justify-content:space-between; font-size:12px; color:var(--mid-grey); margin-bottom:4px; }
.mck-progress-track { background:var(--light-grey); height:6px; width:100%; }
.mck-progress-fill  { background:var(--accent); height:6px; }

/* ── Input overrides ─────────────────────────────────────────────────────────
   Make all Streamlit inputs match the McKinsey off-white / warm-border style. */
[data-testid="stNumberInput"] input,
[data-testid="stTextInput"] input {
    background: var(--off-white) !important; border: 1px solid var(--rule) !important;
    border-radius: 2px !important; color: var(--body) !important;
    font-family: 'Source Code Pro', monospace !important; font-size: 13px !important;
}
[data-testid="stNumberInput"] input:focus,
[data-testid="stTextInput"] input:focus { border-color: var(--accent) !important; box-shadow:none !important; }
[data-testid="stSelectbox"] > div > div {
    background: var(--off-white) !important; border: 1px solid var(--rule) !important;
    border-radius: 2px !important;
}
label { font-size:13px !important; color:#444 !important; font-weight:500 !important; }

/* ── File uploader — styled to match form inputs ────────────────────────────── */
[data-testid="stFileUploader"] {
    border: 2px dashed var(--accent) !important;
    border-radius: 4px !important;
    background: var(--accent-lt) !important;
    padding: 8px !important;
}

/* ── Buttons ─────────────────────────────────────────────────────────────── */
.stButton > button {
    background: var(--navy) !important; color: #FFFFFF !important;
    border: none !important; border-radius: 2px !important;
    font-family: 'Source Sans 3', sans-serif !important;
    font-size:13px !important; font-weight:600 !important;
    text-transform:uppercase !important; letter-spacing:0.06em !important; padding:8px 20px !important;
}
.stButton > button:hover { background: var(--accent) !important; }

/* ── Expanders ─────────────────────────────────────────────────────────────── */
[data-testid="stExpander"] {
    border: 1px solid var(--rule) !important; border-radius: 2px !important;
    background: var(--off-white) !important;
}
[data-testid="stExpander"] summary { font-weight: 600 !important; font-size: 13px !important; color: var(--navy) !important; }

/* ── Download buttons ────────────────────────────────────────────────────── */
[data-testid="stDownloadButton"] button {
    background: transparent !important; border: 1.5px solid var(--navy) !important;
    color: var(--navy) !important; border-radius: 2px !important;
    font-size:12px !important; font-weight:600 !important;
    text-transform:uppercase !important; letter-spacing:0.06em !important;
}
[data-testid="stDownloadButton"] button:hover { background: var(--navy) !important; color: #FFFFFF !important; }

hr { border:none; border-top:1px solid var(--rule) !important; margin:20px 0 !important; }
footer { display:none; }
#MainMenu { display:none; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  SESSION STATE INITIALISATION
#
#  Streamlit reruns the entire script on every interaction. st.session_state
#  persists values across those reruns. init_state() sets safe defaults once;
#  on every subsequent rerun the "if k not in" guard keeps user edits intact.
# ══════════════════════════════════════════════════════════════════════════════
def init_state():
    defaults = {
        # Company context
        "company_name":   "Your Company",
        "industry":       "Technology",
        "country":        "United States",
        "reporting_year": 2025,

        # Financial metrics
        "revenue_musd": 0.0,
        "employees":    0,

        # GWP factors — IPCC AR6 defaults
        "gwp_ch4_fossil": 29.8,
        "gwp_n2o":        273.0,
        "gwp_hfc134a":    1526.0,
        "gwp_sf6":        25200.0,

        # Scope 1 — Stationary Combustion
        "s1_natgas_mmbtu":   0.0,
        "s1_diesel_litres":  0.0,
        "s1_lpg_litres":     0.0,
        "s1_coal_shorttons": 0.0,

        # Scope 1 — Mobile Combustion
        "s1_gasoline_litres":     0.0,
        "s1_diesel_fleet_litres": 0.0,
        "s1_jet_litres":          0.0,

        # Scope 1 — Fugitive Emissions
        "s1_hfc134a_kg": 0.0,
        "s1_hfc410a_kg": 0.0,
        "s1_sf6_kg":     0.0,

        # Scope 2
        "s2_elec_mwh":  0.0,
        "s2_grid_ef":   386.0,
        "s2_market_ef": 0.0,
        "s2_recs_mwh":  0.0,
        "s2_steam_gj":  0.0,

        # Scope 3
        "s3_cat1_spend":      0.0,
        "s3_cat1_ef":         0.35,
        "s3_cat3_elec_mwh":   0.0,
        "s3_cat6_air_km":     0.0,
        "s3_cat6_rail_km":    0.0,
        "s3_cat7_km_per_emp": 0.0,
        "s3_cat11_units":     0.0,
        "s3_cat11_ef":        0.0,

        # Prior year actuals
        "prior_s1":   0.0,
        "prior_s2mb": 0.0,
        "prior_s3":   0.0,

        # Reduction target
        "target_year":          2030,
        "target_reduction_pct": 50.0,
        "target_baseline":      0.0,

        # Industry benchmark
        "benchmark_revenue_intensity": 0.0,

        # Excel import tracking — stores a summary of what was successfully read
        "excel_imported": False,    # True once a file has been parsed
        "excel_filename": "",       # Name of the uploaded file
        "excel_summary":  [],       # List of strings describing what was imported
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


# ══════════════════════════════════════════════════════════════════════════════
#  LOCKED EMISSION FACTORS — EPA / DEFRA / IEA 2023
# ══════════════════════════════════════════════════════════════════════════════
EF = {
    "natgas_co2": 53.06, "natgas_ch4": 1.0,   "natgas_n2o": 0.1,
    "diesel_co2": 2.663, "diesel_ch4": 0.139,  "diesel_n2o": 0.014,
    "lpg_co2":    1.555, "lpg_ch4":    0.0617, "lpg_n2o":    0.0062,
    "coal_co2": 2325.0,  "coal_ch4": 274.0,    "coal_n2o":   40.0,
    "gasoline_co2": 2.312, "gasoline_ch4": 0.339, "gasoline_n2o": 0.033,
    "jet_co2": 2.553,
    "hfc410a_gwp": 2088,
    "s3_cat3_td_loss": 0.05,
    "s3_cat6_air": 0.255, "s3_cat6_rail": 0.041,
    "s3_cat7_car": 0.170,
    "s2_steam_ef": 66.4,
}

COUNTRY_EF = {
    "United States": 386, "United Kingdom": 207, "Germany": 380,
    "France": 85, "China": 581, "India": 713, "Japan": 471,
    "Australia": 680, "Brazil": 119, "Canada": 130, "Other": 386,
}


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL IMPORT ENGINE
#
#  This is the core new feature. It reads the uploaded .xlsx file using
#  openpyxl (data_only=True so we get calculated values, not formula strings),
#  then maps specific cell addresses to session_state variables.
#
#  The cell map below matches the exact structure of Lab1_dashboard.xlsx:
#    - Assumptions sheet: B9 = company name, B15 = revenue, etc.
#    - Scope 1 sheet: C6 = natural gas MMBtu, C21 = diesel fleet litres, etc.
#    - Scope 2 sheet: C23 = total electricity, F23 = RECs, H32 = LB total, etc.
#    - Scope 3 sheet: computed totals from J33, J42, J54, I68, J70
#    - Reduction Trajectory sheet: historical data for prior-year totals
#
#  SAFETY: every read is wrapped in try/except so a missing or non-numeric
#  cell never crashes the app — it just skips that field silently.
# ══════════════════════════════════════════════════════════════════════════════

def safe_num(value, default=0.0):
    """
    Convert a cell value to float safely.
    Excel cells can contain strings, None, or formulas that evaluated to errors.
    This function returns the default if conversion fails.
    """
    try:
        if value is None:
            return default
        f = float(value)
        # Treat Excel error codes (which openpyxl returns as strings like "#REF!")
        # as missing data by checking for the hash sign.
        if str(value).startswith("#"):
            return default
        return f
    except (TypeError, ValueError):
        return default


def safe_str(value, default=""):
    """Return a stripped string from a cell value, or the default if empty."""
    if value is None:
        return default
    s = str(value).strip()
    return s if s else default


def parse_excel(uploaded_file):
    """
    Read the uploaded Excel workbook and populate st.session_state with the
    values found in each sheet. Returns a list of human-readable strings
    describing what was successfully imported (shown in the success banner).

    Steps:
      1. Load the workbook in data_only mode (reads calculated values).
      2. For each sheet that exists, read the relevant input cells.
      3. Write valid values into session_state.
      4. Collect a summary of what was found.

    Args:
        uploaded_file: the BytesIO object returned by st.file_uploader()

    Returns:
        summary (list[str]): description lines for the import banner
    """
    summary = []

    # ── Load the workbook ────────────────────────────────────────────────────
    # io.BytesIO wraps the uploaded bytes so openpyxl can treat it like a file.
    # data_only=True is CRITICAL — without it, openpyxl returns formula strings
    # instead of the computed numbers.
    wb = load_workbook(io.BytesIO(uploaded_file.read()), data_only=True)
    sheet_names = wb.sheetnames  # list of all sheet tab names in the workbook

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 1: ASSUMPTIONS
    #  Cell map — exact addresses from Lab1_dashboard.xlsx:
    #    B9  = company name
    #    B10 = industry / sector
    #    B11 = country of operations
    #    B12 = reporting year (current)
    #    B15 = annual revenue in USD millions
    #    B16 = number of employees (FTE)
    #    E25 = GWP for CH₄ (fossil) — user's selected value
    #    E27 = GWP for N₂O
    #    E28 = GWP for HFC-134a
    #    E29 = GWP for SF₆
    # ══════════════════════════════════════════════════════════════════════════
    if "Assumptions" in sheet_names:
        ws = wb["Assumptions"]

        name = safe_str(ws["B9"].value)
        if name:
            st.session_state["company_name"] = name

        industry = safe_str(ws["B10"].value)
        if industry:
            # Map the Excel industry to the closest option in our dropdown list
            industry_map = {
                "technology": "Technology",
                "transportation": "Transportation & Logistics",
                "logistics": "Transportation & Logistics",
                "manufacturing": "Manufacturing",
                "retail": "Retail & Consumer",
                "financial": "Financial Services",
                "healthcare": "Healthcare & Pharma",
                "energy": "Energy & Utilities",
                "real estate": "Real Estate",
                "food": "Food & Beverage",
            }
            matched = next(
                (v for k, v in industry_map.items() if k in industry.lower()),
                "Other"
            )
            st.session_state["industry"] = matched

        country = safe_str(ws["B11"].value)
        if country:
            # Only accept countries we have grid EF data for
            known = list(COUNTRY_EF.keys())
            matched_country = next(
                (c for c in known if c.lower() in country.lower()), "United States"
            )
            st.session_state["country"] = matched_country

        yr = safe_num(ws["B12"].value, 0)
        if yr > 2000:
            st.session_state["reporting_year"] = int(yr)

        rev = safe_num(ws["B15"].value, 0)
        if rev > 0:
            st.session_state["revenue_musd"] = rev

        emp = safe_num(ws["B16"].value, 0)
        if emp > 0:
            st.session_state["employees"] = int(emp)

        # GWP factors — column E holds "Your Selection" values
        gwp_ch4 = safe_num(ws["E25"].value, 0)
        if gwp_ch4 > 0:
            st.session_state["gwp_ch4_fossil"] = gwp_ch4

        gwp_n2o = safe_num(ws["E27"].value, 0)
        if gwp_n2o > 0:
            st.session_state["gwp_n2o"] = gwp_n2o

        gwp_hfc = safe_num(ws["E28"].value, 0)
        if gwp_hfc > 0:
            st.session_state["gwp_hfc134a"] = gwp_hfc

        gwp_sf6 = safe_num(ws["E29"].value, 0)
        if gwp_sf6 > 0:
            st.session_state["gwp_sf6"] = gwp_sf6

        summary.append(
            f"Company: <strong>{st.session_state['company_name']}</strong> "
            f"| Industry: {st.session_state['industry']} "
            f"| Country: {st.session_state['country']} "
            f"| Year: {st.session_state['reporting_year']}"
        )
        summary.append(
            f"Revenue: <strong>${st.session_state['revenue_musd']:,.0f}M</strong> "
            f"| Employees: <strong>{st.session_state['employees']:,} FTE</strong>"
        )

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 2: SCOPE 1
    #  Activity quantities are in column C, rows 6–34.
    #  The template structure:
    #    C6  = natural gas MMBtu (stationary)
    #    C7  = natural gas m³   (stationary, alternative unit — we skip)
    #    C8  = diesel litres    (stationary boilers)
    #    C10 = LPG litres       (stationary)
    #    C11 = coal short tons  (stationary)
    #    C18 = gasoline litres  (fleet — passenger cars)
    #    C21 = diesel litres    (fleet — heavy-duty trucks)
    #    C23 = aviation fuel litres (company aircraft)
    #    C29 = HFC-134a kg leaked   (fugitive)
    #    C30 = HFC-410A kg leaked   (fugitive)
    #    C31 = SF₆ kg leaked        (fugitive)
    #
    #  K37 = SCOPE 1 TOTAL (tCO₂e) — we read this as the prior-year baseline
    #        candidate if prior year sheets aren't populated.
    # ══════════════════════════════════════════════════════════════════════════
    if "Scope 1" in sheet_names:
        ws1 = wb["Scope 1"]

        ng = safe_num(ws1["C6"].value)
        st.session_state["s1_natgas_mmbtu"] = ng

        # Diesel stationary (row 8, litres)
        ds = safe_num(ws1["C8"].value)
        st.session_state["s1_diesel_litres"] = ds

        # LPG stationary (row 10)
        lpg = safe_num(ws1["C10"].value)
        st.session_state["s1_lpg_litres"] = lpg

        # Coal (row 11, short tons)
        coal = safe_num(ws1["C11"].value)
        st.session_state["s1_coal_shorttons"] = coal

        # Fleet gasoline — passenger cars (row 18)
        gas_fleet = safe_num(ws1["C18"].value)
        st.session_state["s1_gasoline_litres"] = gas_fleet

        # Fleet diesel — heavy-duty trucks (row 21)
        # If row 21 has data, use it. Otherwise try row 20 (diesel passenger cars).
        diesel_fleet = safe_num(ws1["C21"].value)
        if diesel_fleet == 0:
            diesel_fleet = safe_num(ws1["C20"].value)
        st.session_state["s1_diesel_fleet_litres"] = diesel_fleet

        # Aviation fuel (row 23)
        jet = safe_num(ws1["C23"].value)
        st.session_state["s1_jet_litres"] = jet

        # Fugitives (rows 29–31)
        st.session_state["s1_hfc134a_kg"] = safe_num(ws1["C29"].value)
        st.session_state["s1_hfc410a_kg"] = safe_num(ws1["C30"].value)
        st.session_state["s1_sf6_kg"]     = safe_num(ws1["C31"].value)

        # Read the template's Scope 1 total (K37) to use as prior-year reference
        s1_total_xl = safe_num(ws1["K37"].value)

        summary.append(
            f"Scope 1 inputs imported — "
            f"Natural Gas: {ng:,.0f} MMBtu | "
            f"Diesel fleet: {diesel_fleet:,.0f} L | "
            f"Excel total: {s1_total_xl:,.0f} tCO₂e"
        )

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 3: SCOPE 2
    #  The electricity table (rows 16–22) has one row per facility group.
    #  We read the SUBTOTAL row (row 23) which aggregates them all:
    #    C23 = total MWh consumed (SUM of C16:C22)
    #    F23 = total RECs/PPAs covered MWh (SUM of F16:F22)
    #    H23 = location-based total tCO₂e  (SUM of H16:H22)
    #    I23 = market-based total tCO₂e    (SUM of I16:I22)
    #
    #  We also read the average market-based EF from the individual rows:
    #    E16 = market EF for first facility (we use this as representative)
    #
    #  Steam/heat (rows 27–29, column C):
    #    C27 = purchased steam GJ
    #
    #  Totals (bottom of sheet):
    #    H32 = Scope 2 location-based TOTAL
    #    I33 = Scope 2 market-based TOTAL  ← the one we report
    # ══════════════════════════════════════════════════════════════════════════
    if "Scope 2" in sheet_names:
        ws2 = wb["Scope 2"]

        # Total electricity from the subtotal row
        elec_mwh = safe_num(ws2["C23"].value)
        if elec_mwh > 0:
            st.session_state["s2_elec_mwh"] = elec_mwh

        # RECs / PPAs from the subtotal row
        recs = safe_num(ws2["F23"].value)
        if recs >= 0:
            st.session_state["s2_recs_mwh"] = min(recs, elec_mwh)

        # Location-based grid EF: derive from the location-based total and MWh
        lb_total = safe_num(ws2["H32"].value)
        mb_total = safe_num(ws2["I33"].value)
        if elec_mwh > 0 and lb_total > 0:
            # Back-calculate: EF = (tCO₂e × 1000) / MWh  [kg/MWh]
            derived_grid_ef = (lb_total * 1000) / elec_mwh
            st.session_state["s2_grid_ef"] = round(derived_grid_ef, 1)

        # Market-based EF: derive from market-based total and net MWh
        net_mwh = max(0, elec_mwh - recs)
        if net_mwh > 0 and mb_total > 0:
            # steam portion is unknown here, so this is an approximation
            derived_mb_ef = (mb_total * 1000) / net_mwh
            st.session_state["s2_market_ef"] = round(derived_mb_ef, 1)

        # Purchased steam (row 27, column C)
        steam = safe_num(ws2["C27"].value)
        if steam > 0:
            st.session_state["s2_steam_gj"] = steam

        # Store the Excel computed Scope 2 MB total for prior-year comparison
        s2_mb_total_xl = mb_total

        summary.append(
            f"Scope 2 imported — "
            f"Electricity: {elec_mwh:,.0f} MWh | "
            f"RECs/PPAs: {recs:,.0f} MWh | "
            f"Market-Based: {mb_total:,.0f} tCO₂e"
        )

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 4: SCOPE 3
    #  The template breaks Scope 3 into sections:
    #    Category 1 (rows 25–32): individual spend rows, column C = spend USD
    #      J33 = Cat 1 subtotal (tCO₂e)
    #    Category 3 (rows 38–41): T&D losses + WTT fuels
    #      C38 = electricity MWh (pulled from Scope 2)
    #      J42 = Cat 3 subtotal (tCO₂e) — but may be text in some versions
    #    Category 6 (rows 48–53): business travel distances
    #      C48 = air short-haul km
    #      C49 = air long-haul km
    #      C51 = rail km
    #      J54 = Cat 6 subtotal (tCO₂e)
    #    Additional categories (rows 58–66): column I = pre-computed tCO₂e
    #      I68 = additional categories subtotal
    #    J70 = SCOPE 3 GRAND TOTAL
    #
    #  Strategy: we prefer to read the SUBTOTAL tCO₂e values (J33, J42, J54,
    #  I68, J70) directly because they incorporate the Excel formulas.
    #  For our recalculation in the app, we back-populate the activity inputs
    #  from the raw data columns (C column) where readable.
    # ══════════════════════════════════════════════════════════════════════════
    if "Scope 3" in sheet_names:
        ws3 = wb["Scope 3"]

        # ── Category 1: Purchased Goods & Services ─────────────────────────
        # Sum all individual spend rows (C25:C32) to get total annual spend
        cat1_spend = sum(
            safe_num(ws3.cell(row=r, column=3).value)  # column C = column 3
            for r in range(25, 33)
        )
        if cat1_spend > 0:
            st.session_state["s3_cat1_spend"] = cat1_spend / 1000  # convert $ → $000s

        # Use the EEIO factor from first spend row (D25) as representative
        cat1_ef = safe_num(ws3["D25"].value, 0.35)
        if cat1_ef > 0:
            st.session_state["s3_cat1_ef"] = cat1_ef

        # ── Category 3: Fuel & Energy T&D ─────────────────────────────────
        # Row 38, column C = electricity MWh for T&D calculation
        cat3_elec = safe_num(ws3["C38"].value)
        if cat3_elec > 0:
            st.session_state["s3_cat3_elec_mwh"] = cat3_elec
        elif "s2_elec_mwh" in st.session_state and st.session_state["s2_elec_mwh"] > 0:
            # Fall back: use the Scope 2 electricity total
            st.session_state["s3_cat3_elec_mwh"] = st.session_state["s2_elec_mwh"]

        # ── Category 6: Business Travel ────────────────────────────────────
        air_sh = safe_num(ws3["C48"].value)   # short-haul air km
        air_lh = safe_num(ws3["C49"].value)   # long-haul air km
        rail   = safe_num(ws3["C51"].value)   # rail km

        # Combine short-haul and long-haul into a single air total for simplicity
        st.session_state["s3_cat6_air_km"]  = air_sh + air_lh
        st.session_state["s3_cat6_rail_km"] = rail

        # ── Scope 3 grand total from Excel ─────────────────────────────────
        s3_total_xl = safe_num(ws3["J70"].value)

        summary.append(
            f"Scope 3 imported — "
            f"Cat1 spend: ${cat1_spend/1e6:,.1f}M | "
            f"Air travel: {air_sh+air_lh:,.0f} km | "
            f"Excel S3 total: {s3_total_xl:,.0f} tCO₂e"
        )

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 5: REDUCTION TRAJECTORY
    #  This sheet holds the historical (prior year) scope breakdown.
    #  Column F = FY2023 (Year -1), Column G = FY2024 (most recent full year).
    #  Rows:
    #    17 = Scope 1 historical
    #    19 = Scope 2 market-based historical
    #    21 = Scope 3 historical
    #    22 = Total historical
    #
    #  We use the column BEFORE the current year as "prior year".
    #  The sheet has years in row 16: C16=2020, D=2021, E=2022, F=2023, G=2024.
    #  We pick column F (2023) as the prior year since reporting year = 2024/2025.
    # ══════════════════════════════════════════════════════════════════════════
    if "Reduction Trajectory" in sheet_names:
        wrt = wb["Reduction Trajectory"]

        # Column F = year index 6 (A=1, B=2, C=3, D=4, E=5, F=6)
        # These are the FY2023 actuals — one year before the current reporting year
        prior_s1  = safe_num(wrt.cell(row=17, column=6).value)  # F17
        prior_s2mb= safe_num(wrt.cell(row=19, column=6).value)  # F19
        prior_s3  = safe_num(wrt.cell(row=21, column=6).value)  # F21
        prior_tot = safe_num(wrt.cell(row=22, column=6).value)  # F22

        if prior_s1 > 0:
            st.session_state["prior_s1"]   = prior_s1
        if prior_s2mb > 0:
            st.session_state["prior_s2mb"] = prior_s2mb
        if prior_s3 > 0:
            st.session_state["prior_s3"]   = prior_s3

        # Use the earliest year total as the reduction target baseline (column C = 2020)
        baseline = safe_num(wrt.cell(row=22, column=3).value)  # C22
        if baseline > 0:
            st.session_state["target_baseline"] = baseline

        summary.append(
            f"Prior year (FY2023) — "
            f"S1: {prior_s1:,.0f} | S2MB: {prior_s2mb:,.0f} | S3: {prior_s3:,.0f} tCO₂e"
        )

    # ══════════════════════════════════════════════════════════════════════════
    #  SHEET 6: TREND ANALYSIS
    #  Reads the reduction targets the company has committed to:
    #    B18 = target year for Net Zero / primary target
    #    E18 = % reduction required (as a decimal, e.g. 0.30 = 30%)
    # ══════════════════════════════════════════════════════════════════════════
    if "Trend Analysis" in sheet_names:
        wt = wb["Trend Analysis"]

        target_yr = safe_num(wt["B18"].value, 0)
        if target_yr > 2024:
            st.session_state["target_year"] = int(target_yr)

        target_pct = safe_num(wt["E18"].value, 0)
        if 0 < target_pct <= 1:
            # The Excel stores it as a decimal (0.30 = 30%)
            st.session_state["target_reduction_pct"] = target_pct * 100
        elif target_pct > 1:
            # Already in percentage form (30 = 30%)
            st.session_state["target_reduction_pct"] = target_pct

        summary.append(
            f"Reduction target: {st.session_state['target_reduction_pct']:.0f}% "
            f"by {st.session_state['target_year']}"
        )

    return summary


# ══════════════════════════════════════════════════════════════════════════════
#  GHG CALCULATION FUNCTIONS
#  These are pure functions: they only read from session_state and EF,
#  and always return fresh numbers. Keeping them separate from UI code
#  means they run correctly whether values came from manual input or Excel import.
# ══════════════════════════════════════════════════════════════════════════════

def calc_scope1():
    s = st.session_state
    gwp_ch4, gwp_n2o = s["gwp_ch4_fossil"], s["gwp_n2o"]

    # Inner function: convert activity quantity + three EFs → 3-tuple (CO₂, CH₄, N₂O) in tCO₂e
    def tco2e(co2, ch4, n2o, qty):
        return (co2*qty/1000, ch4*qty/1e6*gwp_ch4, n2o*qty/1e6*gwp_n2o)

    r = {}
    # Stationary combustion
    r["natgas"]       = tco2e(EF["natgas_co2"],   EF["natgas_ch4"],   EF["natgas_n2o"],   s["s1_natgas_mmbtu"])
    r["diesel"]       = tco2e(EF["diesel_co2"],   EF["diesel_ch4"],   EF["diesel_n2o"],   s["s1_diesel_litres"])
    r["lpg"]          = tco2e(EF["lpg_co2"],      EF["lpg_ch4"],      EF["lpg_n2o"],      s["s1_lpg_litres"])
    r["coal"]         = tco2e(EF["coal_co2"],     EF["coal_ch4"],     EF["coal_n2o"],     s["s1_coal_shorttons"])
    # Mobile combustion
    r["gasoline"]     = tco2e(EF["gasoline_co2"], EF["gasoline_ch4"], EF["gasoline_n2o"], s["s1_gasoline_litres"])
    r["diesel_fleet"] = tco2e(EF["diesel_co2"],   EF["diesel_ch4"],   EF["diesel_n2o"],   s["s1_diesel_fleet_litres"])
    r["jet"]          = (EF["jet_co2"] * s["s1_jet_litres"] / 1000, 0, 0)
    # Fugitives (GWP already embedded)
    fug = (s["s1_hfc134a_kg"]*s["gwp_hfc134a"]/1000
         + s["s1_hfc410a_kg"]*EF["hfc410a_gwp"]/1000
         + s["s1_sf6_kg"]*s["gwp_sf6"]/1000)
    r["fugitive"] = (fug, 0, 0)
    return r, sum(sum(v) for v in r.values())


def calc_scope2():
    s = st.session_state
    net = max(0, s["s2_elec_mwh"] - s["s2_recs_mwh"])
    steam = s["s2_steam_gj"] * EF["s2_steam_ef"] / 1000
    lb = s["s2_elec_mwh"] * s["s2_grid_ef"] / 1000 + steam
    mb = net * s["s2_market_ef"] / 1000 + steam
    recs_pct = s["s2_recs_mwh"] / max(1, s["s2_elec_mwh"]) * 100
    return {"lb": lb, "mb": mb, "net": net, "recs_pct": recs_pct}


def calc_scope3():
    s = st.session_state
    cat1  = s["s3_cat1_spend"] * s["s3_cat1_ef"] / 1000
    cat3  = s["s3_cat3_elec_mwh"] * s["s2_grid_ef"] * EF["s3_cat3_td_loss"] / 1000
    cat6  = (s["s3_cat6_air_km"]*EF["s3_cat6_air"] + s["s3_cat6_rail_km"]*EF["s3_cat6_rail"]) / 1000
    cat7  = s["s3_cat7_km_per_emp"] * s["employees"] * EF["s3_cat7_car"] / 1000
    cat11 = s["s3_cat11_units"] * s["s3_cat11_ef"] / 1000
    return {"cat1":cat1,"cat3":cat3,"cat6":cat6,"cat7":cat7,"cat11":cat11,
            "total":cat1+cat3+cat6+cat7+cat11}


# ── Formatting helpers ─────────────────────────────────────────────────────────
def fi(v):   return f"{v:,.0f}"    # integer with thousands comma
def ff(v):   return f"{v:,.1f}"    # 1 decimal with thousands comma
def dpct(c, p): return (c-p)/p*100 if p > 0 else None  # % change vs prior year


# ── UI component helpers ───────────────────────────────────────────────────────
def kpi_html(label, value, unit, delta=None):
    """Build an HTML KPI card string. delta<0 = teal (good), delta>0 = amber (bad)."""
    d = ""
    if delta is not None:
        cls = "delta-neg" if delta < 0 else "delta-pos"
        arr = "↓" if delta < 0 else "↑"
        d = f'<div class="mck-kpi-delta {cls}">{arr} {abs(delta):.1f}% vs prior year</div>'
    return f"""<div class="mck-kpi">
      <div class="mck-kpi-label">{label}</div>
      <div class="mck-kpi-value">{value}</div>
      <div class="mck-kpi-unit">{unit}</div>{d}
    </div>"""


def section_head(eyebrow, title, caption=""):
    """Render the standard McKinsey page header: eyebrow → h1 → caption."""
    st.markdown(f'<span class="eyebrow">{eyebrow}</span>', unsafe_allow_html=True)
    st.markdown(f"# {title}")
    if caption:
        st.markdown(f'<p class="page-caption">{caption}</p>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR
#  Contains: branding, Excel upload widget, navigation, live totals.
#  The upload widget is in the sidebar so it's always accessible from any page.
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:

    # ── Illinois Tech branding block ──────────────────────────────────────────
    # Rendered as pure HTML/CSS text — no image file, no CDN, no black box.
    # Matches the screenshot exactly:
    #   - "ILLINOIS TECH" in Illinois Tech scarlet (#CC0000), bold, large
    #   - "SAM 503 ESG Analytics & Management" in white, smaller, below
    #   - Both sit flush on the navy sidebar — transparent background
    #
    # Why text instead of image:
    #   PNG images carry their own background colour (black in this case).
    #   There is no CSS way to remove a PNG background without transparency.
    #   Using styled HTML text gives us pixel-perfect control and zero
    #   dependencies — no file needed, works on Streamlit Cloud instantly.

    st.markdown(
        '<div style="padding:20px 0 16px 0;">'

        # ① "ILLINOIS TECH" — scarlet, heavy weight, large tracking
        # font-family uses a web-safe condensed-style fallback stack that
        # approximates the Illinois Tech typeface without any web font import.
        '<div style="'
        'font-family: Arial Black, Arial, Helvetica, sans-serif;'
        'font-weight: 900;'
        'font-size: 1.55rem;'
        'letter-spacing: 0.04em;'
        'color: #D12030;'          # Illinois Tech scarlet
        'line-height: 1.1;'
        'margin-bottom: 5px;'
        '">'
        'ILLINOIS TECH'
        '</div>'

        # ② "SAM 503 ESG Analytics & Management" — white, normal weight, small
        '<div style="'
        'font-family: Source Sans 3, Arial, sans-serif;'
        'font-weight: 400;'
        'font-size: 11px;'
        'letter-spacing: 0.03em;'
        'color: #FFFFFF;'
        'margin-bottom: 14px;'
        '">'
        'SAM 503 ESG Analytics &amp; Management'
        '</div>'

        # ③ Thin separator between the IIT brand and the app label
        '<div style="border-top:1px solid rgba(255,255,255,0.18);'
        'margin-bottom:12px;"></div>'

        # ④ App title
        '<div style="'
        'font-family: Playfair Display, Georgia, serif;'
        'font-size: 1.1rem;'
        'font-weight: 700;'
        'color: #FFFFFF;'
        'letter-spacing: -0.01em;'
        'line-height: 1.2;'
        'margin-bottom: 4px;'
        '">'
        'GHG Carbon Inventory'
        '</div>'

        # ⑤ Framework sub-label
        '<div style="'
        'font-size: 10px;'
        'color: #7DA3CC;'
        'text-transform: uppercase;'
        'letter-spacing: 0.12em;'
        '">'
        'GHG Protocol Framework'
        '</div>'

        '</div>',
        unsafe_allow_html=True,
    )

    st.divider()

    # ── Excel Upload Widget     
    
    
    # ── Excel Upload Widget ───────────────────────────────────────────────────
    # This is the heart of the new feature.
    # st.file_uploader() returns None until the user selects a file,
    # then returns a UploadedFile object (which behaves like a BytesIO stream).
    #
    # We compare the uploaded filename against the last-imported filename
    # stored in session_state to detect when a NEW file has been dropped —
    # this prevents re-importing the same file on every Streamlit rerun.
    st.markdown(
        '<div style="font-size:10px;text-transform:uppercase;letter-spacing:0.12em;'
        'color:#7DA3CC;margin-bottom:8px;">Import Excel</div>',
        unsafe_allow_html=True,
    )

    uploaded = st.file_uploader(
        "Upload company Excel",        # visible label (hidden by CSS but kept for a11y)
        type=["xlsx"],                  # only accept Excel files
        label_visibility="collapsed",   # hide the label text in the UI
        help="Upload your Lab1_dashboard.xlsx (or compatible GHG template) "
             "to auto-populate all fields.",
    )

    # ── Process the uploaded file ─────────────────────────────────────────────
    if uploaded is not None:
        # Check if this is a new file (different name from last import)
        # OR if a file has never been imported yet.
        is_new_file = (uploaded.name != st.session_state.get("excel_filename", ""))

        if is_new_file:
            # Show a spinner while parsing — large workbooks can take a moment
            with st.spinner("Reading Excel…"):
                try:
                    summary_lines = parse_excel(uploaded)
                    # Mark import as successful in session_state
                    st.session_state["excel_imported"]  = True
                    st.session_state["excel_filename"]  = uploaded.name
                    st.session_state["excel_summary"]   = summary_lines
                except Exception as e:
                    # If parsing fails entirely, show an error but don't crash
                    st.session_state["excel_imported"] = False
                    st.error(f"Could not read Excel: {e}")

        # Show a compact success badge once a file has been imported
        if st.session_state["excel_imported"]:
            st.markdown(
                f'<div style="background:rgba(26,110,60,0.2);border-left:3px solid #1A6E3C;'
                f'padding:8px 10px;border-radius:2px;margin-top:6px;">'
                f'<div style="font-size:10px;font-weight:600;text-transform:uppercase;'
                f'letter-spacing:0.1em;color:#4CAF82;">✓ Imported</div>'
                f'<div style="font-size:11px;color:#A8D5BC;margin-top:2px;">'
                f'{st.session_state["excel_filename"]}</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

    # Small instruction when no file has been uploaded yet
    elif not st.session_state.get("excel_imported", False):
        st.markdown(
            '<div style="font-size:11px;color:#7DA3CC;margin-top:4px;">'
            'Drop your GHG template .xlsx here to auto-fill all fields.'
            '</div>',
            unsafe_allow_html=True,
        )

    st.divider()

    # ── Navigation ────────────────────────────────────────────────────────────
    page = st.radio("Nav", [
        "Assumptions",
        "Scope 1 — Direct",
        "Scope 2 — Purchased Energy",
        "Scope 3 — Value Chain",
        "Dashboard",
        "Export Report",
    ], label_visibility="collapsed")

    st.divider()

    # ── Live totals ───────────────────────────────────────────────────────────
    # Recalculate on every rerun so the sidebar always reflects current values.
    _, s1t = calc_scope1()
    s2     = calc_scope2()
    s3     = calc_scope3()
    grand  = s1t + s2["mb"] + s3["total"]

    st.markdown(
        '<div style="font-size:10px;text-transform:uppercase;letter-spacing:0.12em;'
        'color:#7DA3CC;margin-bottom:8px;">Live Totals</div>',
        unsafe_allow_html=True,
    )
    st.metric("Scope 1",      f"{fi(s1t)} tCO₂e")
    st.metric("Scope 2 (MB)", f"{fi(s2['mb'])} tCO₂e")
    st.metric("Scope 3",      f"{fi(s3['total'])} tCO₂e")
    st.metric("Total",        f"{fi(grand)} tCO₂e")

    # Company name + year at the bottom of the sidebar
    st.markdown(
        f'<div style="margin-top:20px;padding-top:16px;'
        f'border-top:1px solid rgba(255,255,255,0.12);font-size:11px;color:#7DA3CC;">'
        f'{st.session_state["company_name"]}<br>FY {st.session_state["reporting_year"]}</div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  IMPORT SUCCESS BANNER
#  Shown at the top of every page immediately after a successful Excel import.
#  Displays what was read from each sheet so the user can verify correctness.
#  The banner auto-disappears once the user edits any field manually.
# ══════════════════════════════════════════════════════════════════════════════
# Import summary is shown only on the Export Report page (see PAGE 6)


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE ROUTING
#  Streamlit has no built-in router. We chain if/elif on the "page" variable
#  set by the sidebar radio. Only the matching block renders.
# ══════════════════════════════════════════════════════════════════════════════


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE 1: ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════════════
if page == "Assumptions":
    section_head(
        "Step 01", "Central Assumptions",
        "All calculation sheets reference these inputs. "
        "Upload an Excel file on the left to auto-populate, or fill in manually.",
    )

    col1, col2 = st.columns(2, gap="large")

    with col1:
        st.markdown("## Company Profile")
        st.session_state["company_name"] = st.text_input(
            "Legal Entity / Company Name", st.session_state["company_name"])
        st.session_state["industry"] = st.selectbox("Industry Sector", [
            "Technology", "Manufacturing", "Retail & Consumer", "Financial Services",
            "Healthcare & Pharma", "Energy & Utilities", "Real Estate",
            "Transportation & Logistics", "Food & Beverage", "Other",
        ], index=max(0, ["Technology","Manufacturing","Retail & Consumer","Financial Services",
            "Healthcare & Pharma","Energy & Utilities","Real Estate",
            "Transportation & Logistics","Food & Beverage","Other"
        ].index(st.session_state["industry"]) if st.session_state["industry"] in [
            "Technology","Manufacturing","Retail & Consumer","Financial Services",
            "Healthcare & Pharma","Energy & Utilities","Real Estate",
            "Transportation & Logistics","Food & Beverage","Other"] else 0))
        st.session_state["country"] = st.selectbox(
            "Primary Country of Operations", list(COUNTRY_EF.keys()),
            index=list(COUNTRY_EF.keys()).index(st.session_state["country"])
                  if st.session_state["country"] in COUNTRY_EF else 0)
        st.session_state["reporting_year"] = st.number_input(
            "Reporting Year", 2000, 2030, st.session_state["reporting_year"])

    with col2:
        st.markdown("## Financial Metrics")
        st.caption("Used to calculate carbon intensity ratios.")
        st.session_state["revenue_musd"] = st.number_input(
            "Annual Revenue (USD Millions)", 0.0,
            value=float(st.session_state["revenue_musd"]), step=100.0)
        st.session_state["employees"] = st.number_input(
            "Full-Time Equivalent Employees (FTE)", 0,
            value=int(st.session_state["employees"]), step=100)

        st.markdown("## Prior Year Actuals")
        st.caption("Used for year-on-year comparison. Auto-populated from Reduction Trajectory sheet if available.")
        pa, pb, pc = st.columns(3)
        with pa:
            st.session_state["prior_s1"] = st.number_input(
                "Scope 1 (tCO₂e)", 0.0, value=float(st.session_state["prior_s1"]))
        with pb:
            st.session_state["prior_s2mb"] = st.number_input(
                "Scope 2 MB (tCO₂e)", 0.0, value=float(st.session_state["prior_s2mb"]))
        with pc:
            st.session_state["prior_s3"] = st.number_input(
                "Scope 3 (tCO₂e)", 0.0, value=float(st.session_state["prior_s3"]))

    st.divider()
    st.markdown("## GWP Factors — IPCC AR6 (2021)")
    st.caption("Auto-populated from Assumptions!E25–E29 when Excel is imported. "
               "Adjust only if your regulatory regime requires AR5 or AR4.")
    gc1, gc2, gc3, gc4 = st.columns(4)
    with gc1:
        st.session_state["gwp_ch4_fossil"] = st.number_input(
            "CH₄ fossil GWP", value=float(st.session_state["gwp_ch4_fossil"]))
        st.markdown('<span class="ef-chip">AR5: 25 | AR6: 29.8</span>', unsafe_allow_html=True)
    with gc2:
        st.session_state["gwp_n2o"] = st.number_input(
            "N₂O GWP", value=float(st.session_state["gwp_n2o"]))
        st.markdown('<span class="ef-chip">AR5: 298 | AR6: 273</span>', unsafe_allow_html=True)
    with gc3:
        st.session_state["gwp_hfc134a"] = st.number_input(
            "HFC-134a GWP", value=float(st.session_state["gwp_hfc134a"]))
        st.markdown('<span class="ef-chip">AR5: 1430 | AR6: 1526</span>', unsafe_allow_html=True)
    with gc4:
        st.session_state["gwp_sf6"] = st.number_input(
            "SF₆ GWP", value=float(st.session_state["gwp_sf6"]))
        st.markdown('<span class="ef-chip">AR5: 22800 | AR6: 25200</span>', unsafe_allow_html=True)

    st.divider()
    st.markdown("## Reduction Target")
    st.caption("Auto-populated from Trend Analysis sheet if present.")
    ta, tb, tc = st.columns(3)
    with ta:
        st.session_state["target_year"] = st.number_input(
            "Target Year", 2025, 2060, int(st.session_state["target_year"]))
    with tb:
        st.session_state["target_reduction_pct"] = st.slider(
            "Reduction vs Baseline (%)", 0, 100, int(st.session_state["target_reduction_pct"]))
    with tc:
        st.session_state["target_baseline"] = st.number_input(
            "Baseline Emissions (tCO₂e)", 0.0, value=float(st.session_state["target_baseline"]))

    if st.session_state["target_baseline"] > 0:
        target_val = st.session_state["target_baseline"] * (
            1 - st.session_state["target_reduction_pct"] / 100)
        yrs = max(1, st.session_state["target_year"] - st.session_state["reporting_year"])
        annual = (st.session_state["target_baseline"] - target_val) / yrs
        st.markdown(
            f'<div class="insight-box"><strong>Target trajectory:</strong> '
            f'Reduce to <strong>{fi(target_val)} tCO₂e</strong> by '
            f'{st.session_state["target_year"]}, requiring '
            f'<strong>{fi(annual)} tCO₂e/year</strong> over {yrs} years.</div>',
            unsafe_allow_html=True,
        )

    st.session_state["benchmark_revenue_intensity"] = st.number_input(
        "Industry Benchmark — Revenue Intensity (tCO₂e / $M revenue)", 0.0,
        value=float(st.session_state["benchmark_revenue_intensity"]))


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE 2: SCOPE 1 — DIRECT EMISSIONS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Scope 1 — Direct":
    results, s1t = calc_scope1()
    stationary = sum(sum(results[k]) for k in ["natgas","diesel","lpg","coal"])
    mobile     = sum(sum(results[k]) for k in ["gasoline","diesel_fleet","jet"])
    fugitive   = sum(results["fugitive"])

    section_head(
        "Step 02", "Scope 1 — Direct GHG Emissions",
        "Sources directly owned or controlled. "
        "Auto-populated from Scope 1 sheet columns C (activity quantities).",
    )

    # Import note when data came from Excel
    if st.session_state.get("excel_imported"):
        st.markdown(
            '<div class="insight-box">Fields below were auto-populated from the Excel upload. '
            'You can edit any value and the totals will update instantly.</div>',
            unsafe_allow_html=True,
        )

    c1, c2, c3, c4 = st.columns(4)
    with c1: st.markdown(kpi_html("Total Scope 1",         fi(s1t),        "tCO₂e"), unsafe_allow_html=True)
    with c2: st.markdown(kpi_html("Stationary Combustion", fi(stationary), "tCO₂e"), unsafe_allow_html=True)
    with c3: st.markdown(kpi_html("Mobile Combustion",     fi(mobile),     "tCO₂e"), unsafe_allow_html=True)
    with c4: st.markdown(kpi_html("Fugitive Emissions",    fi(fugitive),   "tCO₂e"), unsafe_allow_html=True)

    st.divider()
    st.markdown("## Part A — Stationary Combustion")
    st.caption("Source: EPA CCCL 2023 | Excel: Scope 1!C6, C8, C10, C11")

    for label, key, unit, ef_val, rkey in [
        ("Natural Gas",             "s1_natgas_mmbtu",   "MMBtu",      EF["natgas_co2"],    "natgas"),
        ("Diesel / Fuel Oil No.2",  "s1_diesel_litres",  "Litres",     EF["diesel_co2"],    "diesel"),
        ("Liquefied Petroleum Gas", "s1_lpg_litres",     "Litres",     EF["lpg_co2"],       "lpg"),
        ("Coal (Bituminous)",       "s1_coal_shorttons", "Short Tons", EF["coal_co2"]/1000, "coal"),
    ]:
        row_total = sum(results[rkey])
        with st.expander(f"{label}   —   {ff(row_total)} tCO₂e"):
            ea, eb, ec_ = st.columns([3,2,2])
            with ea:
                st.session_state[key] = st.number_input(
                    f"Activity Quantity ({unit})", 0.0,
                    value=float(st.session_state[key]),
                    key=f"si_{key}",
                    step=1.0 if "ton" in unit.lower() else 100.0)
            with eb:
                st.markdown(f"<br><span class='ef-chip'>CO₂ EF: {ef_val:.4f} kgCO₂/{unit}</span>", unsafe_allow_html=True)
            with ec_:
                st.metric("Result", f"{ff(row_total)} tCO₂e")

    st.divider()
    st.markdown("## Part B — Mobile Combustion")
    st.caption("Source: DEFRA 2023 / EPA 2023 | Excel: Scope 1!C18, C21, C23")
    mb1, mb2, mb3 = st.columns(3)
    with mb1:
        st.session_state["s1_gasoline_litres"] = st.number_input(
            "Gasoline / Petrol (Litres)", 0.0, value=float(st.session_state["s1_gasoline_litres"]), step=100.0)
        st.markdown(f"<span class='ef-chip'>{EF['gasoline_co2']} kgCO₂/L</span>", unsafe_allow_html=True)
    with mb2:
        st.session_state["s1_diesel_fleet_litres"] = st.number_input(
            "Diesel Fleet (Litres)", 0.0, value=float(st.session_state["s1_diesel_fleet_litres"]), step=100.0)
        st.markdown(f"<span class='ef-chip'>{EF['diesel_co2']} kgCO₂/L</span>", unsafe_allow_html=True)
    with mb3:
        st.session_state["s1_jet_litres"] = st.number_input(
            "Aviation Fuel / Jet-A (Litres)", 0.0, value=float(st.session_state["s1_jet_litres"]), step=100.0)
        st.markdown(f"<span class='ef-chip'>{EF['jet_co2']} kgCO₂/L</span>", unsafe_allow_html=True)

    st.divider()
    st.markdown("## Part C — Fugitive Emissions")
    st.caption("Source: IPCC AR6 GWP | Excel: Scope 1!C29, C30, C31")
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        st.session_state["s1_hfc134a_kg"] = st.number_input(
            "HFC-134a leaked (kg)", 0.0, value=float(st.session_state["s1_hfc134a_kg"]))
        st.markdown(f"<span class='ef-chip'>GWP: {st.session_state['gwp_hfc134a']}</span>", unsafe_allow_html=True)
    with fc2:
        st.session_state["s1_hfc410a_kg"] = st.number_input(
            "HFC-410A leaked (kg)", 0.0, value=float(st.session_state["s1_hfc410a_kg"]))
        st.markdown(f"<span class='ef-chip'>GWP: {EF['hfc410a_gwp']}</span>", unsafe_allow_html=True)
    with fc3:
        st.session_state["s1_sf6_kg"] = st.number_input(
            "SF₆ leaked (kg)", 0.0, value=float(st.session_state["s1_sf6_kg"]))
        st.markdown(f"<span class='ef-chip'>GWP: {st.session_state['gwp_sf6']}</span>", unsafe_allow_html=True)

    st.divider()
    st.markdown("## Scope 1 Summary")
    rows = [
        ("Natural Gas",       "Stationary", ff(sum(results["natgas"]))),
        ("Diesel / Fuel Oil", "Stationary", ff(sum(results["diesel"]))),
        ("LPG",               "Stationary", ff(sum(results["lpg"]))),
        ("Coal",              "Stationary", ff(sum(results["coal"]))),
        ("Gasoline Fleet",    "Mobile",     ff(sum(results["gasoline"]))),
        ("Diesel Fleet",      "Mobile",     ff(sum(results["diesel_fleet"]))),
        ("Aviation",          "Mobile",     ff(sum(results["jet"]))),
        ("Fugitive",          "Fugitive",   ff(sum(results["fugitive"]))),
    ]
    tbl = ('<table class="mck-table"><thead><tr>'
           '<th>Source</th><th>Category</th><th class="num">tCO₂e</th>'
           '</tr></thead><tbody>')
    for n, c, v in rows:
        tbl += f"<tr><td>{n}</td><td>{c}</td><td class='num'>{v}</td></tr>"
    tbl += (f'<tr class="total-row"><td><strong>Total Scope 1</strong></td><td></td>'
            f'<td class="num"><strong>{ff(s1t)}</strong></td></tr></tbody></table>')
    st.markdown(tbl, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE 3: SCOPE 2 — PURCHASED ENERGY
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Scope 2 — Purchased Energy":
    s2 = calc_scope2()

    section_head(
        "Step 03", "Scope 2 — Purchased Energy",
        "GHG Protocol dual reporting: Market-Based (primary) and Location-Based (supplemental). "
        "Auto-populated from Scope 2!C23, F23, H32, I33.",
    )

    if st.session_state.get("excel_imported"):
        st.markdown(
            '<div class="insight-box">Electricity total, RECs/PPAs, and derived emission factors '
            'were read from the Excel subtotal rows. Review and adjust as needed.</div>',
            unsafe_allow_html=True,
        )

    c1, c2, c3 = st.columns(3)
    with c1: st.markdown(kpi_html("Market-Based (Primary)",  fi(s2["mb"]),       "tCO₂e"), unsafe_allow_html=True)
    with c2: st.markdown(kpi_html("Location-Based (Suppl.)", fi(s2["lb"]),       "tCO₂e"), unsafe_allow_html=True)
    with c3: st.markdown(kpi_html("Renewable Coverage",      f"{s2['recs_pct']:.1f}%", "of total electricity"), unsafe_allow_html=True)

    st.divider()
    st.markdown("## Part A — Purchased Electricity")
    ea, eb, ec = st.columns(3)
    with ea:
        st.session_state["s2_elec_mwh"] = st.number_input(
            "Total Electricity Consumed (MWh)", 0.0,
            value=float(st.session_state["s2_elec_mwh"]), step=100.0)
    with eb:
        suggested = COUNTRY_EF.get(st.session_state["country"], 386)
        st.session_state["s2_grid_ef"] = st.number_input(
            "Location-Based Grid EF (kgCO₂e/MWh)", 0.0,
            value=float(st.session_state["s2_grid_ef"]), step=1.0)
        st.markdown(f"<span class='ef-chip'>Country default: {suggested}</span>", unsafe_allow_html=True)
    with ec:
        st.session_state["s2_market_ef"] = st.number_input(
            "Market-Based Supplier EF (kgCO₂e/MWh)", 0.0,
            value=float(st.session_state["s2_market_ef"]), step=1.0)

    st.session_state["s2_recs_mwh"] = st.slider(
        "RECs / PPAs Covered (MWh)", 0.0,
        max(float(st.session_state["s2_elec_mwh"]), 1.0),
        min(float(st.session_state["s2_recs_mwh"]), max(float(st.session_state["s2_elec_mwh"]), 0.0)),
        step=100.0,
    )

    st.markdown(
        f'<div class="insight-box"><strong>Net metered electricity:</strong> '
        f'{fi(s2["net"])} MWh after {fi(st.session_state["s2_recs_mwh"])} MWh RECs/PPAs. '
        f'Market-based emissions = net MWh × {st.session_state["s2_market_ef"]} kgCO₂e/MWh.</div>',
        unsafe_allow_html=True,
    )

    st.divider()
    st.markdown("## Part B — Purchased Steam, Heat & Cooling")
    st.caption("Excel: Scope 2!C27")
    st.session_state["s2_steam_gj"] = st.number_input(
        "Purchased Steam / Heat / Cooling (GJ)", 0.0,
        value=float(st.session_state["s2_steam_gj"]))
    st.markdown(f"<span class='ef-chip'>EF: {EF['s2_steam_ef']} kgCO₂e/GJ</span>", unsafe_allow_html=True)

    st.divider()
    tbl = ('<table class="mck-table"><thead><tr>'
           '<th>Method</th><th>Basis</th><th class="num">tCO₂e</th>'
           '</tr></thead><tbody>')
    tbl += f"<tr><td>Location-Based</td><td>Grid average × total MWh</td><td class='num'>{ff(s2['lb'])}</td></tr>"
    tbl += f"<tr><td>Market-Based</td><td>Supplier EF × net MWh after RECs</td><td class='num'>{ff(s2['mb'])}</td></tr>"
    tbl += (f'<tr class="total-row"><td><strong>Primary (Market-Based)</strong></td>'
            f'<td>Used in total inventory</td><td class="num"><strong>{ff(s2["mb"])}</strong></td>'
            f'</tr></tbody></table>')
    st.markdown(tbl, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE 4: SCOPE 3 — VALUE CHAIN
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Scope 3 — Value Chain":
    s3 = calc_scope3()

    section_head(
        "Step 04", "Scope 3 — Value Chain Emissions",
        "Indirect upstream and downstream emissions. Typically 70–90% of total footprint. "
        "Activity data auto-populated from Scope 3 sheet.",
    )

    if st.session_state.get("excel_imported"):
        st.markdown(
            '<div class="insight-box">Spend data (Cat 1), travel distances (Cat 6), and '
            'electricity for T&D (Cat 3) were read from the Excel. '
            'Cat 7 and Cat 11 require manual entry unless provided.</div>',
            unsafe_allow_html=True,
        )

    c1, _ = st.columns([1, 3])
    with c1: st.markdown(kpi_html("Scope 3 Total", fi(s3["total"]), "tCO₂e"), unsafe_allow_html=True)

    st.divider()

    configs = [
        ("Cat. 1",  "Purchased Goods & Services",        "s3_cat1_spend",      "s3_cat1_ef",
         "Annual Spend ($000s USD)", "EEIO Factor (kgCO₂e per $)",
         "Spend-based | US EPA USEEIO v2.0 | Excel: Scope 3!C25:C32 (summed)", "cat1", True),
        ("Cat. 3",  "Fuel & Energy-Related (T&D Losses)", "s3_cat3_elec_mwh",   None,
         "Electricity consumed (MWh)", None,
         f"5% T&D loss × grid EF ({st.session_state['s2_grid_ef']} kgCO₂e/MWh) | Excel: Scope 3!C38",
         "cat3", False),
        ("Cat. 6",  "Business Travel",                    "s3_cat6_air_km",     None,
         "Total Air Travel (km — short + long haul combined)", None,
         f"Economy class | EF: {EF['s3_cat6_air']} kgCO₂e/km | Excel: Scope 3!C48+C49",
         "cat6", False),
        ("Cat. 7",  "Employee Commuting",                 "s3_cat7_km_per_emp", None,
         "Avg commute per employee/year (km)", None,
         f"Applied to {st.session_state['employees']:,} FTEs | EF: {EF['s3_cat7_car']} kgCO₂e/km | Manual entry",
         "cat7", False),
        ("Cat. 11", "Use of Sold Products",               "s3_cat11_units",     "s3_cat11_ef",
         "Units / product-years in use", "Lifecycle EF (kgCO₂e/unit)",
         "Activity-based | Provide product lifecycle EF | Manual entry", "cat11", True),
    ]

    for cat_num, cat_name, k1, k2, lbl1, lbl2, note, rkey, has_two in configs:
        val = s3[rkey]
        with st.expander(f"**{cat_num} — {cat_name}**   —   {ff(val)} tCO₂e"):
            if has_two:
                ca, cb, cc = st.columns([3,2,2])
                with ca:
                    st.session_state[k1] = st.number_input(
                        lbl1, 0.0, value=float(st.session_state[k1]),
                        step=1000.0, key=f"s3a_{k1}")
                with cb:
                    st.session_state[k2] = st.number_input(
                        lbl2, 0.0, value=float(st.session_state[k2]),
                        step=0.01, key=f"s3b_{k2}")
                with cc:
                    st.metric("Result", f"{ff(val)} tCO₂e")
            else:
                ca, cb = st.columns([3,2])
                with ca:
                    st.session_state[k1] = st.number_input(
                        lbl1, 0.0, value=float(st.session_state[k1]),
                        step=1000.0, key=f"s3a_{k1}")
                with cb:
                    st.metric("Result", f"{ff(val)} tCO₂e")
            st.caption(f"Methodology: {note}")

    st.divider()
    st.markdown("## Category Summary")
    cat_labels = [
        ("cat1",  "Cat 1 — Purchased Goods & Services", "Spend-based"),
        ("cat3",  "Cat 3 — Fuel & Energy (T&D)",        "Activity-based"),
        ("cat6",  "Cat 6 — Business Travel",             "Activity-based"),
        ("cat7",  "Cat 7 — Employee Commuting",          "Activity-based"),
        ("cat11", "Cat 11 — Use of Sold Products",       "Activity-based"),
    ]
    tbl = ('<table class="mck-table"><thead><tr>'
           '<th>Category</th><th>Method</th><th class="num">tCO₂e</th><th class="num">% of S3</th>'
           '</tr></thead><tbody>')
    for k, label, method in cat_labels:
        pct = s3[k] / max(s3["total"], 1) * 100
        tbl += f"<tr><td>{label}</td><td>{method}</td><td class='num'>{ff(s3[k])}</td><td class='num'>{pct:.1f}%</td></tr>"
    tbl += (f'<tr class="total-row"><td><strong>Total Scope 3</strong></td><td></td>'
            f'<td class="num"><strong>{ff(s3["total"])}</strong></td>'
            f'<td class="num"><strong>100%</strong></td></tr></tbody></table>')
    st.markdown(tbl, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE 5: DASHBOARD — EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Dashboard":
    _, s1t = calc_scope1()
    s2     = calc_scope2()
    s3     = calc_scope3()
    grand  = s1t + s2["mb"] + s3["total"]
    prior_grand = (st.session_state["prior_s1"]
                 + st.session_state["prior_s2mb"]
                 + st.session_state["prior_s3"])
    s = st.session_state

    section_head(
        "Executive View",
        f"{s['company_name']} — GHG Inventory {s['reporting_year']}",
        "GHG Protocol Corporate Standard  ·  Market-Based Scope 2  ·  Figures in tCO₂e",
    )

    # Import source badge — shows which file the data came from
    if s.get("excel_imported"):
        st.markdown(
            f'<div style="display:inline-block;background:var(--green-lt);'
            f'border:1px solid var(--green);border-radius:2px;'
            f'padding:4px 12px;font-size:11px;font-weight:600;'
            f'text-transform:uppercase;letter-spacing:0.08em;color:var(--green);'
            f'margin-bottom:16px;">'
            f'◈ Data source: {s["excel_filename"]}</div>',
            unsafe_allow_html=True,
        )

    # ── Top KPI row ───────────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    with c1: st.markdown(kpi_html("Scope 1 — Direct",       fi(s1t),        "tCO₂e", dpct(s1t,        s["prior_s1"])),    unsafe_allow_html=True)
    with c2: st.markdown(kpi_html("Scope 2 — Market-Based", fi(s2["mb"]),   "tCO₂e", dpct(s2["mb"],   s["prior_s2mb"])), unsafe_allow_html=True)
    with c3: st.markdown(kpi_html("Scope 3 — Value Chain",  fi(s3["total"]),"tCO₂e", dpct(s3["total"],s["prior_s3"])),   unsafe_allow_html=True)
    with c4: st.markdown(kpi_html("Total (S1 + S2 + S3)",   fi(grand),      "tCO₂e", dpct(grand, prior_grand)),           unsafe_allow_html=True)

    st.divider()
    col_l, col_r = st.columns([3,2], gap="large")

    with col_l:
        st.markdown("## Scope Breakdown")
        scope_rows = [
            ("Scope 1",      "Direct — combustion, fleet, fugitive",  s1t,        s["prior_s1"]),
            ("Scope 2 (MB)", "Purchased energy — market-based",       s2["mb"],   s["prior_s2mb"]),
            ("Scope 3",      "Value chain — upstream & downstream",   s3["total"],s["prior_s3"]),
            ("Total",        "S1 + S2(MB) + S3",                      grand,      prior_grand),
        ]
        tbl = ('<table class="mck-table"><thead><tr>'
               '<th>Scope</th><th>Description</th>'
               '<th class="num">tCO₂e</th><th class="num">% Share</th><th class="num">YoY</th>'
               '</tr></thead><tbody>')
        for sc, desc, val, prior in scope_rows:
            pct = f"{val/max(grand,1)*100:.1f}%" if sc != "Total" else "100%"
            yoy = f"{dpct(val,prior):+.1f}%" if prior > 0 else "—"
            cls = ' class="total-row"' if sc == "Total" else ""
            b  = "<strong>" if sc == "Total" else ""
            be = "</strong>" if sc == "Total" else ""
            tbl += (f"<tr{cls}><td>{b}{sc}{be}</td><td>{desc}</td>"
                    f"<td class='num'>{b}{fi(val)}{be}</td>"
                    f"<td class='num'>{pct}</td><td class='num'>{yoy}</td></tr>")
        tbl += "</tbody></table>"
        st.markdown(tbl, unsafe_allow_html=True)
        st.markdown(
            f"<div style='font-size:11px;color:var(--mid-grey);margin-top:6px;'>"
            f"◦ Scope 2 Location-Based (supplemental): {fi(s2['lb'])} tCO₂e"
            f"</div>", unsafe_allow_html=True,
        )

    with col_r:
        st.markdown("## Carbon Intensity")
        rev, emp = s["revenue_musd"], s["employees"]
        bench   = s["benchmark_revenue_intensity"]
        tbl2 = ('<table class="mck-table"><thead><tr>'
                '<th>Metric</th><th class="num">Value</th><th class="num">Unit</th><th class="num">YoY</th>'
                '</tr></thead><tbody>')
        if rev > 0:
            ri = grand / rev
            pri_ri = prior_grand / rev if prior_grand > 0 else None
            tbl2 += (f"<tr><td>Revenue intensity</td><td class='num'>{ri:.2f}</td>"
                     f"<td class='num'>tCO₂e/$M</td>"
                     f"<td class='num'>{f'{dpct(ri,pri_ri):+.1f}%' if pri_ri else '—'}</td></tr>")
            if bench > 0:
                vs = (ri - bench) / bench * 100
                tbl2 += (f"<tr><td>vs Benchmark</td><td class='num'>{vs:+.1f}%</td>"
                         f"<td class='num'>Bench: {bench:.1f}</td><td class='num'>—</td></tr>")
        if emp > 0:
            ei = grand / emp
            pri_ei = prior_grand / emp if prior_grand > 0 else None
            tbl2 += (f"<tr><td>Employee intensity</td><td class='num'>{ei:.2f}</td>"
                     f"<td class='num'>tCO₂e/FTE</td>"
                     f"<td class='num'>{f'{dpct(ei,pri_ei):+.1f}%' if pri_ei else '—'}</td></tr>")
        tbl2 += "</tbody></table>"
        if rev > 0 or emp > 0:
            st.markdown(tbl2, unsafe_allow_html=True)
        else:
            st.caption("Enter revenue and employees in Assumptions to see intensity ratios.")

        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("## Renewable Energy")
        rp = s2["recs_pct"]
        st.markdown(
            f'<div style="margin:12px 0;">'
            f'<div class="mck-progress-label"><span>Renewable Coverage</span><span>{rp:.1f}%</span></div>'
            f'<div class="mck-progress-track"><div class="mck-progress-fill" style="width:{min(rp,100):.1f}%"></div></div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # ── Reduction target progress ──────────────────────────────────────────────
    if s["target_baseline"] > 0:
        st.divider()
        st.markdown("## Reduction Target Progress")
        target_val = s["target_baseline"] * (1 - s["target_reduction_pct"] / 100)
        progress   = max(0, min(1, (s["target_baseline"]-grand) / max(1, s["target_baseline"]-target_val)))
        yrs        = max(1, s["target_year"] - s["reporting_year"])
        annual_req = max(0, (grand - target_val) / yrs)

        ta, tb, tc, td = st.columns(4)
        with ta: st.markdown(kpi_html("Baseline",              fi(s["target_baseline"]), "tCO₂e"), unsafe_allow_html=True)
        with tb: st.markdown(kpi_html("Current",               fi(grand),                "tCO₂e"), unsafe_allow_html=True)
        with tc: st.markdown(kpi_html("Target",                fi(target_val),           f"tCO₂e by {s['target_year']}"), unsafe_allow_html=True)
        with td: st.markdown(kpi_html("Annual Reduction Req.", fi(annual_req),           "tCO₂e/year"), unsafe_allow_html=True)

        st.markdown(
            f'<div style="margin-top:16px;">'
            f'<div class="mck-progress-label">'
            f'<span>Progress toward {s["target_reduction_pct"]:.0f}% reduction by {s["target_year"]}</span>'
            f'<span>{progress*100:.0f}%</span></div>'
            f'<div class="mck-progress-track">'
            f'<div class="mck-progress-fill" style="width:{min(progress*100,100):.1f}%"></div>'
            f'</div></div>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE 6: EXPORT REPORT
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Export Report":
    _, s1t = calc_scope1()
    s2     = calc_scope2()
    s3     = calc_scope3()
    grand  = s1t + s2["mb"] + s3["total"]
    s      = st.session_state

    section_head(
        "Step 05", "Export Report",
        "Download formatted outputs or save and restore your complete input dataset.",
    )

    # ── Excel Import Summary ───────────────────────────────────────────────────
    # Only rendered on this page, and only when a file has been successfully
    # imported. Presents each import line exactly as the user requested:
    # bullet-point format with the filename as the header.
    if s.get("excel_imported") and s.get("excel_summary"):

        # Build bullet rows from the summary list stored during parse_excel().
        # Each entry in excel_summary is a plain string describing one sheet's data.
        # We prefix every line with "* " to match the requested format.
        bullet_rows = "".join(
            f'<div style="display:flex;gap:8px;padding:3px 0;">'
            f'<span style="color:var(--accent);font-weight:700;flex-shrink:0;">*</span>'
            f'<span>{ln}</span></div>'
            for ln in s["excel_summary"]
        )

        # Reduction target line — built live from session_state so it always
        # reflects the current target year and percentage, not a stale snapshot.
        target_line = (
            f'<div style="display:flex;gap:8px;padding:3px 0;">'
            f'<span style="color:var(--accent);font-weight:700;flex-shrink:0;">*</span>'
            f'<span>Reduction target: <strong>{s["target_reduction_pct"]:.0f}%</strong>'
            f' by <strong>{s["target_year"]}</strong></span></div>'
        )

        st.markdown(
            f'<div class="success-box">'
            # Header line: checkmark + filename
            f'<div style="font-weight:700;font-size:14px;color:var(--green);margin-bottom:8px;">'
            f'✓ Excel imported: {s["excel_filename"]}</div>'
            # All per-sheet bullet rows
            f'<div style="font-size:13px;line-height:1.7;color:var(--body);">'
            f'{bullet_rows}'
            f'{target_line}'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        st.divider()

    # Note the data source in the report text file
    source_note = f"Data source: {s['excel_filename']}" if s.get("excel_imported") else "Data source: Manual entry"

    report_text = f"""{'═'*65}
  GHG CARBON FOOTPRINT REPORT
  {s['company_name']}  |  FY {s['reporting_year']}
  Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}
  Framework: GHG Protocol Corporate Standard
  {source_note}
{'═'*65}

SECTION 1 — SCOPE TOTALS
{'─'*65}
  Scope 1   Direct Emissions                  {s1t:>13,.1f}  tCO₂e
  Scope 2   Market-Based (Primary)            {s2['mb']:>13,.1f}  tCO₂e
  Scope 2   Location-Based (Supplemental)     {s2['lb']:>13,.1f}  tCO₂e
  Scope 3   Value Chain Emissions             {s3['total']:>13,.1f}  tCO₂e
{'─'*65}
  TOTAL     S1 + S2(MB) + S3                  {grand:>13,.1f}  tCO₂e

SECTION 2 — SCOPE 3 DETAIL
{'─'*65}
  Cat  1   Purchased Goods & Services         {s3['cat1']:>13,.1f}  tCO₂e
  Cat  3   Fuel & Energy (T&D Losses)         {s3['cat3']:>13,.1f}  tCO₂e
  Cat  6   Business Travel                    {s3['cat6']:>13,.1f}  tCO₂e
  Cat  7   Employee Commuting                 {s3['cat7']:>13,.1f}  tCO₂e
  Cat 11   Use of Sold Products               {s3['cat11']:>13,.1f}  tCO₂e

SECTION 3 — CARBON INTENSITY
{'─'*65}"""

    rev, emp = s["revenue_musd"], s["employees"]
    if rev > 0:
        report_text += f"\n  Revenue Intensity          {grand/rev:>10.2f}  tCO₂e / $M revenue"
    if emp > 0:
        report_text += f"\n  Employee Intensity         {grand/emp:>10.2f}  tCO₂e / FTE"

    report_text += f"""

SECTION 4 — ASSUMPTIONS
{'─'*65}
  Revenue                    ${rev:>12,.0f}M USD
  Employees                  {emp:>12,} FTE
  Grid EF (location)         {s['s2_grid_ef']:>12} kgCO₂e/MWh
  Market EF                  {s['s2_market_ef']:>12} kgCO₂e/MWh
  GWP CH₄ (fossil)           {s['gwp_ch4_fossil']:>12}  [IPCC AR6]
  GWP N₂O                    {s['gwp_n2o']:>12}  [IPCC AR6]
  GWP HFC-134a               {s['gwp_hfc134a']:>12}  [IPCC AR6]
  GWP SF₆                    {s['gwp_sf6']:>12}  [IPCC AR6]

SECTION 5 — REDUCTION TARGET
{'─'*65}
  Baseline                   {s['target_baseline']:>12,.0f}  tCO₂e
  Target Year                {s['target_year']:>12}
  Reduction Required         {s['target_reduction_pct']:>11.0f}%
  Target Value               {s['target_baseline']*(1-s['target_reduction_pct']/100):>12,.0f}  tCO₂e

METHODOLOGY
{'─'*65}
  Emission Factors: EPA CCCL 2023, DEFRA 2023, IEA 2023,
    US EPA USEEIO v2.0
  GWP Reference: IPCC AR6, 2021
  Scope 2 Primary Method: Market-Based
  Scope 3 Methods: Spend-based (Cat 1), Activity-based (Cat 3, 6, 7, 11)
{'═'*65}
"""

    d1, d2, d3 = st.columns(3)
    with d1:
        st.download_button(
            "⬇  Report (.txt)", data=report_text,
            file_name=f"GHG_Report_{s['company_name'].replace(' ','_')}_{s['reporting_year']}.txt",
            mime="text/plain", use_container_width=True)
    with d2:
        csv_rows = [["Metric","Value","Unit"],
                    ["Company", s["company_name"],""],["Year",s["reporting_year"],""],
                    ["Scope 1",f"{s1t:.2f}","tCO₂e"],["Scope 2 MB",f"{s2['mb']:.2f}","tCO₂e"],
                    ["Scope 2 LB",f"{s2['lb']:.2f}","tCO₂e"],["Scope 3",f"{s3['total']:.2f}","tCO₂e"],
                    ["Total",f"{grand:.2f}","tCO₂e"]]
        if rev > 0: csv_rows.append(["Revenue Intensity",f"{grand/rev:.4f}","tCO₂e/$M"])
        if emp > 0: csv_rows.append(["Employee Intensity",f"{grand/emp:.4f}","tCO₂e/FTE"])
        st.download_button(
            "⬇  Summary (.csv)",
            data="\n".join(",".join(str(c) for c in r) for r in csv_rows),
            file_name=f"GHG_Summary_{s['company_name'].replace(' ','_')}_{s['reporting_year']}.csv",
            mime="text/csv", use_container_width=True)
    with d3:
        save_keys = [k for k in st.session_state
                     if not k.startswith("si_") and not k.startswith("s3a_") and not k.startswith("s3b_")]
        st.download_button(
            "💾  Save Inputs (.json)",
            data=json.dumps({k: st.session_state[k] for k in save_keys}, indent=2),
            file_name=f"GHG_Inputs_{s['company_name'].replace(' ','_')}.json",
            mime="application/json", use_container_width=True)

    st.divider()
    with st.expander("Report Preview"):
        st.text(report_text)

    st.divider()
    st.markdown("## Restore Saved Inputs")
    st.caption("Upload a previously saved .json file to reload all inputs.")
    uploaded_json = st.file_uploader("Upload JSON", type="json", label_visibility="collapsed")
    if uploaded_json:
        for k, v in json.load(uploaded_json).items():
            st.session_state[k] = v
        st.success("Inputs restored. Navigate to any section to review.")
